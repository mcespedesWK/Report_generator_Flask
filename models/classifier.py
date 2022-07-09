import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

class Classifier():
    # Recibo el nombre del reporte como parametro
    def setUp(this):
        #---------------------#
        #       WORK BOOK     #
        #---------------------#
        wb = load_workbook("test.xlsx")

        #----------------------#
        #       WORK SHEET     #
        #----------------------#
        # grab the active worksheet
        # This will create the active sheet on this work BOOK
        ws = wb.active
        #Provide le sheet with a name
        ws.title = name
        # With teh work book object we call the create fucntion
        ws_1 = wb.create_sheet("Data")

        return ws_1

    def training(X,Y):
        #--------- TRAIN and TEST Models ------------

        # Creo variables para guardar los resultados de la division de mis datos
        # Se divide en 75% training y 25% testing
        # Esta funcion la importo desde la libreria de SCIKIT_LEARN
        # El random_state igual a 0 significa que no habran valores al azar
        X_train, X_test, y_train, y_test = train_test_split(X,Y,random_state= 0)

        return X_train, X_test, y_train, y_test

    # Recibo los datos que ingresa el usuario
    def desicionClassifier(X_train, X_test, y_train, y_test):
        #--------------------------------------#
        # ---     1- DesicionClasifier     -----
        #--------------------------------------#
        # Para saber que libreria utilizar es necesario saber ante que problema nos enfrenteamos
        # En este caso utilizamos un algoritmo de      ----> Clasificacion Multivariable <-------
        # En donde la salida es de 0,1,2
        # 1- Tipo de clasificador
        # Primero creo el objeto
        clf = DecisionTreeClassifier()

        # Le inyecto los datos de entrenamiento
        #2- Entreno con los datos que indico entre parentesis
        clf.fit(X_train,y_train)

        # Aqui es donde tengo que cambiar y enlugar de inyectarle
        #los datos de X_test le inyecto los datos del usuario
        #3- Realizo la prediccion
        # Recordar que el X_test es el valor que el usuario ingresa por pantalla
        y_pred = clf.predict(X_test)

        # Imprimo esa prediccion
        return y_pred

    def randomForest():
        X_train, X_test, y_train, y_test = training()
        # Primero creo el objeto
        # 1- Tipo de clasificador
        clf = RandomForestClassifier()

        # Le inyecto los datos de entrenamiento
        # Entreno con los datos que indico entre parentesis
        clf.fit(X_train, y_train)

        # Realizo la prediccion
        y_pred = clf.predict(X_test)

        y_pred

        # Esto lo que nos dice es el nivel de acierto
        # El porcentaje de datos que han sido predecidos correctamente
        acc = accuracy_score(y_test, y_pred)
